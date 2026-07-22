"""
IIT Madras Labs & Equipment Browser — Regeneration Script
=========================================================
Run this script whenever the source Excel file is updated.
It reads the data, cleans it, and regenerates the HTML browser.

Usage:
    python3 regenerate_browser.py

Requirements:
    pip install openpyxl

Place this script in the same folder as your Excel source file.
Update SOURCE_FILE below if the filename changes.

CHANGE HISTORY (all changes are labelled in the HTML template below):
  [CHANGE 1] Header logos — replaces the placeholder gold circle with real images.
             Place iitm_logo.png and iitmrp_logo.png next to the output HTML file.
  [CHANGE 2] Equipment-name-only search bar added next to the global search bar.
             Styled with a gold-tinted icon to distinguish it. Filters only
             the 'equipment' field; global search still searches all fields.
  [CHANGE 3] "All Laboratories" filter dropdown COMMENTED OUT.
             To re-enable: search for "CHANGE 3 RE-ENABLE" throughout this file
             and uncomment the marked sections, then restore the <select> in HTML.
  [CHANGE 4] Table view bug fixed — card-view now hides properly when table is
             selected. Sticky table header implemented: sticks just below the
             sticky page header while scrolling down.
  [CHANGE 5] AI Equipment Assistant — a chat widget (floating button, bottom
             right) that lets a user describe a research need in plain
             language and get ranked equipment/lab matches. Calls a small
             serverless proxy (Cloudflare Worker, see worker.js) which holds
             the Gemini API key server-side — no secret is ever embedded in
             this generated HTML, so it is safe to commit to a public repo
             / GitHub Pages. Configure ASSISTANT_PROXY_URL below to point at
             your deployed Worker. See DEPLOYMENT.md for the one-time setup.
"""

import json
import os
import openpyxl

# ── Configuration ─────────────────────────────────────────────────────────────
SOURCE_FILE = "IIT-M_L&E_Web&Visit_Data_OG.xlsx"
OUTPUT_FILE = "index.html"

# ── [CHANGE 5] AI Equipment Assistant configuration ─────────────────────────
# This page never holds a Gemini API key — it calls a small serverless proxy
# (Cloudflare Worker) that holds the key server-side. See DEPLOYMENT.md for
# the one-time setup (free, ~10 minutes, no credit card). Once deployed,
# paste the Worker's URL below — that URL is NOT a secret and is safe to
# commit to a public repo.
#
# Leave ASSISTANT_PROXY_URL empty to ship the page with the assistant
# disabled (it will show a friendly "not configured" message instead of
# erroring).
ASSISTANT_PROXY_URL = "https://iitm-equipment-assistant.gemini-ai-studio-api.workers.dev"

DEPT_MAP = {
    'Department of Physics': 'Physics',
    'Physics': 'Physics',
    'Ocean Engineering': 'Ocean Engineering',
    'Ocean engineering': 'Ocean Engineering',
    'Metallurgical and Materials': 'Metallurgical & Materials Engineering',
    'Metallurgical and Materials Engineering': 'Metallurgical & Materials Engineering',
    'Metallurgical and materials engineering': 'Metallurgical & Materials Engineering',
    'Medical Science and Technology': 'Medical Science & Technology',
    'Mechanical Engineering': 'Mechanical Engineering',
    'Mechanical engineering': 'Mechanical Engineering',
    'Aerospace engineering': 'Aerospace Engineering',
    'Applied mechanics': 'Applied Mechanics',
    'Applied mechanics & biomedical engineering': 'Applied Mechanics',
    'Applied mechanics and biomedical engineering': 'Applied Mechanics',
    'Civil engineering': 'Civil Engineering',
    'Electrical Engineering': 'Electrical Engineering',
    'Electrical engineering': 'Electrical Engineering',
    'Engineering Design': 'Engineering Design',
    'Engineering design': 'Engineering Design',
    'Bio-technology': 'Biotechnology',
    'Biotechnology': 'Biotechnology',
    'Chemical Engineering': 'Chemical Engineering',
    'Chemical engineering': 'Chemical Engineering',
    'Chemistry': 'Chemistry',
    'Ic & sr': 'IC & SR',
    'ARCI': 'ARCI',
}

SHEET_TO_DEPT = {
    'Department of Physics': 'Physics',
    'Ocean Engineering': 'Ocean Engineering',
    'Metallurgical and Materials': 'Metallurgical & Materials Engineering',
    'Medical Science and Technology': 'Medical Science & Technology',
    'Mechanical Engineering': 'Mechanical Engineering',
    'Aerospace engineering': 'Aerospace Engineering',
    'Applied mechanics': 'Applied Mechanics',
    'Civil engineering': 'Civil Engineering',
    'Electrical Engineering': 'Electrical Engineering',
    'Engineering Design': 'Engineering Design',
    'Bio-Technology': 'Biotechnology',
    'Chemical Engineering': 'Chemical Engineering',
    'Department of Chemistry': 'Chemistry',
    'Humanities and Social Sciences': 'Humanities & Social Sciences',
    'Computer Science & Engineering': 'Computer Science & Engineering',
    'DOMS (Manegerial Studies)': 'Management Studies',
    'Department of Mathematics': 'Mathematics',
}

DEPT_SHEETS = list(SHEET_TO_DEPT.keys())


def clean(v):
    if v is None:
        return ''
    return str(v).strip().replace('\n', ' ').replace('\xa0', '').strip()


def extract_data(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True)
    equipment = []

    for sheet_name in DEPT_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [SKIP] Sheet not found: {sheet_name}")
            continue
        ws = wb[sheet_name]
        fallback_dept = SHEET_TO_DEPT[sheet_name]
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not row[0] or not row[1]:
                continue
            lab   = clean(row[0])
            equip = clean(row[1])
            prof  = clean(row[2])
            prof_email = clean(row[3])
            raw_dept   = clean(row[4])
            operator   = clean(row[5])
            op_email   = clean(row[6])
            contact    = clean(row[7])
            source     = clean(row[8]) if len(row) > 8 else ''
            dept = DEPT_MAP.get(raw_dept, '') or fallback_dept
            if lab and equip:
                equipment.append({
                    'lab': lab, 'equipment': equip, 'professor': prof,
                    'prof_email': prof_email, 'department': dept,
                    'operator': operator, 'op_email': op_email,
                    'contact': contact, 'source': source,
                })

    # ERS / ICSR sheet
    if 'ERS_List_ICSR provided data' in wb.sheetnames:
        ws_ers = wb['ERS_List_ICSR provided data']
        first = True
        for row in ws_ers.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not row[0] or not row[1]:
                continue
            lab   = clean(row[0])
            equip = clean(row[1])
            prof  = clean(row[2])
            dept  = DEPT_MAP.get(clean(row[3]), clean(row[3]) or 'IC & SR')
            operator = clean(row[4])
            op_email = clean(row[5])
            contact  = clean(row[6])
            if lab and equip:
                equipment.append({
                    'lab': lab, 'equipment': equip, 'professor': prof,
                    'prof_email': '', 'department': dept,
                    'operator': operator, 'op_email': op_email,
                    'contact': contact, 'source': 'ERS/ICSR',
                })

    # ARCI sheet
    if 'ARCI' in wb.sheetnames:
        ws_arci = wb['ARCI']
        first = True
        for row in ws_arci.iter_rows(values_only=True):
            if first:
                first = False
                continue
            if not row or not row[0] or not row[1]:
                continue
            lab   = clean(row[0])
            equip = clean(row[1])
            centre = clean(row[2])
            if lab and equip:
                equipment.append({
                    'lab': lab, 'equipment': equip, 'professor': centre,
                    'prof_email': '', 'department': 'ARCI',
                    'operator': '', 'op_email': '', 'contact': '', 'source': 'ARCI',
                })

    print(f"  Extracted {len(equipment)} equipment entries from {len(wb.sheetnames)} sheets")
    return equipment


HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>1. IIT Madras/IITM Research Park Ecosystem(Beta) — Labs & Equipments</title>
<style>
  :root {
    --navy: #0A2747;
    --navy-mid: #0D3461;
    --gold: #C9953A;
    --gold-light: #E8B96A;
    --cream: #F7F5F0;
    --white: #FFFFFF;
    --text: #1C2B3A;
    --text-muted: #5A6A7A;
    --border: #DDD8CF;
    --card-bg: #FFFFFF;
    --tag-bg: #EEF3FA;
    --tag-text: #2A4A7F;
    --hover: #F0EDE8;
    --shadow: 0 2px 8px rgba(10,39,71,0.10);
    --shadow-lg: 0 8px 32px rgba(10,39,71,0.14);
    --radius: 10px;
    --radius-sm: 6px;
  }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', system-ui, sans-serif; background: var(--cream); color: var(--text); min-height: 100vh; font-size: 14px; line-height: 1.5; }
  header { background: linear-gradient(135deg, var(--navy) 0%, var(--navy-mid) 100%); color: white; position: sticky; top: 0; z-index: 100; box-shadow: 0 2px 16px rgba(0,0,0,0.25); }
  .header-top { display: flex; align-items: center; gap: 18px; padding: 16px 28px; border-bottom: 1px solid rgba(201,149,58,0.30); }
  .logo-block { display: flex; align-items: center; gap: 14px; flex-shrink: 0; }

  /* ── CHANGE 1: Institution logos ─────────────────────────────────────────────
     .logos-wrap holds both logo images side-by-side with a gold separator line.
     Place iitm_logo.png and iitmrp_logo.png in the same folder as the HTML.     */
  .logos-wrap { display: flex; align-items: center; gap: 10px; border-right: 1px solid rgba(201,149,58,0.35); padding-right: 14px; }
  .logo-img { height: 48px; width: auto; flex-shrink: 0; object-fit: contain; }
  /* ── END CHANGE 1 ────────────────────────────────────────────────────────── */

  .logo-text h1 { font-size: 17px; font-weight: 700; color: white; }
  .logo-text p { font-size: 11px; color: rgba(255,255,255,0.65); letter-spacing: 0.8px; text-transform: uppercase; }
  .header-stats { display: flex; gap: 24px; margin-left: auto; }
  .stat-pill { display: flex; flex-direction: column; align-items: center; background: rgba(255,255,255,0.08); border: 1px solid rgba(255,255,255,0.12); border-radius: 8px; padding: 6px 16px; }
  .stat-pill .num { font-size: 20px; font-weight: 800; color: var(--gold-light); }
  .stat-pill .lbl { font-size: 10px; text-transform: uppercase; letter-spacing: 0.8px; color: rgba(255,255,255,0.55); }
  /* [CHANGE 6] Info bar — contact guidance shown between header-top and filter-bar */
  .header-info-bar { padding: 8px 28px; font-size: 13px; color: rgba(255,255,255,0.88); background: rgba(0,0,0,0.15); border-top: 1px solid rgba(201,149,58,0.20); border-bottom: 1px solid rgba(201,149,58,0.20); line-height: 1.55; }
  .header-info-bar a { color: var(--gold-light); text-decoration: underline; text-underline-offset: 2px; }
  .filter-bar { display: flex; align-items: center; gap: 10px; padding: 12px 28px; flex-wrap: wrap; background: rgba(0,0,0,0.12); }

  /* Shared search-wrap styles (used by both global and equipment-only search) */
  .search-wrap { position: relative; flex: 1; min-width: 200px; }
  .search-wrap svg { position: absolute; left: 12px; top: 50%; transform: translateY(-50%); color: rgba(255,255,255,0.5); pointer-events: none; }
  .search-wrap input { width: 100%; background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; padding: 9px 12px 9px 38px; color: white; font-size: 13px; outline: none; }
  .search-wrap input::placeholder { color: rgba(255,255,255,0.45); }
  .search-wrap input:focus { border-color: var(--gold-light); background: rgba(255,255,255,0.15); }

  /* ── CHANGE 2: Equipment-name-only search bar — accent styling ───────────────
     Uses a gold-tinted icon and border to visually distinguish it from the
     global search. Filters ONLY the equipment name field (not lab/prof/dept).  */
  .equip-only-wrap { flex: 0.85; min-width: 185px; }
  .equip-only-wrap input { border-color: rgba(201,149,58,0.40); }
  .equip-only-wrap input:focus { border-color: var(--gold-light); background: rgba(255,255,255,0.15); }
  .equip-only-wrap svg { color: rgba(201,149,58,0.80); }
  /* ── END CHANGE 2 ────────────────────────────────────────────────────────── */

  select { background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; padding: 9px 32px 9px 12px; color: white; font-size: 12px; outline: none; cursor: pointer; appearance: none; -webkit-appearance: none; background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='12' viewBox='0 0 24 24'%3E%3Cpath fill='rgba(255,255,255,0.5)' d='M7 10l5 5 5-5z'/%3E%3C/svg%3E"); background-repeat: no-repeat; background-position: right 10px center; min-width: 160px; }
  select option { background: var(--navy); color: white; }
  .view-toggle { display: flex; background: rgba(255,255,255,0.10); border: 1px solid rgba(255,255,255,0.18); border-radius: 8px; overflow: hidden; }
  .view-btn { background: none; border: none; color: rgba(255,255,255,0.55); padding: 8px 12px; cursor: pointer; transition: all 0.2s; display: flex; align-items: center; }
  .view-btn.active { background: var(--gold); color: white; }
  .clear-btn { background: none; border: 1px solid rgba(255,255,255,0.20); color: rgba(255,255,255,0.65); border-radius: 8px; padding: 8px 14px; cursor: pointer; font-size: 12px; white-space: nowrap; }
  .clear-btn:hover { background: rgba(255,255,255,0.10); color: white; }
  main { padding: 20px 28px; }
  .results-meta { display: flex; align-items: center; margin-bottom: 16px; }
  .results-count { font-size: 13px; color: var(--text-muted); }
  .results-count strong { color: var(--navy); font-weight: 700; }

  /* ── CHANGE 4 (CSS part 1): Card-view hidden by default ─────────────────────
     BUG FIX: The original code had no display:none on #card-view, so it stayed
     visible even when table view was active. Now hidden unless .active is set.  */
  #card-view { display: none; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 14px; }
  #card-view.active { display: grid; }
  /* ── END CHANGE 4 (CSS part 1) ───────────────────────────────────────────── */

  .eq-card { background: var(--card-bg); border: 1px solid var(--border); border-radius: var(--radius); padding: 16px; box-shadow: var(--shadow); transition: transform 0.15s, box-shadow 0.15s; position: relative; overflow: hidden; }
  .eq-card::before { content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px; background: var(--dept-color, var(--gold)); }
  .eq-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-lg); }
  .card-dept-tag { display: inline-block; font-size: 10px; font-weight: 600; letter-spacing: 0.6px; text-transform: uppercase; background: var(--tag-bg); color: var(--tag-text); border-radius: 4px; padding: 2px 8px; margin-bottom: 8px; }
  .card-equip-name { font-size: 15px; font-weight: 700; color: var(--navy); margin-bottom: 4px; line-height: 1.3; }
  .card-lab-name { font-size: 12px; color: var(--text-muted); margin-bottom: 10px; display: flex; align-items: flex-start; gap: 5px; }
  .card-lab-name svg { flex-shrink: 0; margin-top: 1px; }
  .card-divider { height: 1px; background: var(--border); margin: 10px 0; }
  .card-info-row { display: flex; align-items: flex-start; gap: 8px; font-size: 12px; color: var(--text-muted); margin-top: 5px; }
  .card-info-row svg { flex-shrink: 0; margin-top: 1px; }
  .card-info-row a { color: var(--gold); text-decoration: none; }
  .card-info-row .label { font-weight: 600; color: var(--text); min-width: 65px; }
 #table-view { display: none; }
  .table-scroll { overflow: auto; max-height: calc(100vh - var(--header-h, 130px) - 60px); border-radius: var(--radius); box-shadow: var(--shadow); }
  #table-view.active { display: block; }
  table { width: 100%; border-collapse: collapse; background: white; font-size: 13px; }
  thead { background: var(--navy); color: white; }
  th { padding: 12px 14px; text-align: left; font-size: 11px; font-weight: 600; letter-spacing: 0.7px; text-transform: uppercase; white-space: nowrap; cursor: pointer; user-select: none; }

  /* ── CHANGE 4 (CSS part 2): Sticky table header ──────────────────────────────
     thead th sticks just below the sticky page header when scrolling down.
     --header-h is a CSS variable set dynamically by JS (see bottom of <script>).
     The z-index:90 keeps table header below the page header (z-index:100).
     background:var(--navy) prevents the table body from showing through.        */
  thead th { position: sticky; top: 0; z-index: 90; background: var(--navy); }
  /* ── END CHANGE 4 (CSS part 2) ───────────────────────────────────────────── */

  tbody tr { border-bottom: 1px solid var(--border); transition: background 0.1s; }
  tbody tr:hover { background: var(--hover); }
  td { padding: 11px 14px; vertical-align: top; }
  td:first-child { font-weight: 600; color: var(--navy); }
  .dept-badge { display: inline-block; font-size: 10px; font-weight: 600; border-radius: 4px; padding: 2px 7px; white-space: nowrap; }
  .contact-link { color: var(--gold); text-decoration: none; font-size: 12px; }
  .dept-Physics{--dept-color:#8B5CF6}.dept-Aerospace-Engineering{--dept-color:#0EA5E9}.dept-Mechanical-Engineering{--dept-color:#F59E0B}.dept-Civil-Engineering{--dept-color:#10B981}.dept-Electrical-Engineering{--dept-color:#EF4444}.dept-Chemical-Engineering{--dept-color:#EC4899}.dept-Biotechnology{--dept-color:#14B8A6}.dept-Chemistry{--dept-color:#84CC16}.dept-Ocean-Engineering{--dept-color:#0284C7}.dept-Metallurgical{--dept-color:#78716C}.dept-Applied-Mechanics{--dept-color:#F97316}.dept-ARCI{--dept-color:#A855F7}.dept-IC-SR{--dept-color:#1D4ED8}.dept-Engineering-Design{--dept-color:#D946EF}.dept-default{--dept-color:var(--gold)}
  .badge-Physics{background:#EDE9FE;color:#5B21B6}.badge-Aerospace{background:#E0F2FE;color:#0369A1}.badge-Mechanical{background:#FEF3C7;color:#92400E}.badge-Civil{background:#D1FAE5;color:#065F46}.badge-Electrical{background:#FEE2E2;color:#991B1B}.badge-Chemical{background:#FCE7F3;color:#9D174D}.badge-Biotechnology{background:#CCFBF1;color:#0F766E}.badge-Chemistry{background:#ECFCCB;color:#365314}.badge-Ocean{background:#E0F2FE;color:#075985}.badge-Metal{background:#F5F5F4;color:#44403C}.badge-Applied{background:#FFEDD5;color:#9A3412}.badge-ARCI{background:#F3E8FF;color:#6B21A8}.badge-ICSR{background:#DBEAFE;color:#1E3A8A}.badge-ED{background:#FDF4FF;color:#86198F}.badge-default{background:#F3F4F6;color:#374151}
  .empty-state { text-align: center; padding: 80px 20px; color: var(--text-muted); }
  .empty-state h3 { font-size: 18px; color: var(--text); margin-bottom: 6px; }
  footer { text-align: center; padding: 24px; font-size: 12px; color: var(--text-muted); border-top: 1px solid var(--border); margin-top: 20px; }
  @media (max-width: 700px) { .header-stats{display:none} .header-top,.filter-bar{padding:12px 16px} main{padding:14px 16px} #card-view{grid-template-columns:1fr} }

  .header-info-bar a,
.header-info-bar .link-color {
  color: #E8B96A; /* use your existing link color */
  font-weight: bold;
}

  /* ════════════════════════════════════════════════════════════════════════
     [CHANGE 5] AI Equipment Assistant — chat widget styling
     ════════════════════════════════════════════════════════════════════════ */
  #ai-fab { position: fixed; right: 24px; bottom: 24px; width: 56px; height: 56px; border-radius: 50%; background: linear-gradient(135deg, var(--gold) 0%, var(--gold-light) 100%); border: none; box-shadow: var(--shadow-lg); cursor: pointer; display: flex; align-items: center; justify-content: center; color: var(--navy); z-index: 300; transition: transform 0.15s; }
  #ai-fab:hover { transform: scale(1.06); }
  #ai-fab svg { width: 24px; height: 24px; }
  /* [CHANGE 6] Backdrop — always in DOM; fades in/out via opacity+visibility.
     visibility delay on close keeps it non-interactive only after fade completes. */
  #ai-backdrop { position: fixed; inset: 0; background: rgba(0,0,0,0.50); z-index: 250;
    visibility: hidden; opacity: 0; pointer-events: none;
    transition: opacity 0.25s ease, visibility 0s linear 0.25s; }
  #ai-backdrop.show { visibility: visible; opacity: 1; pointer-events: auto;
    transition: opacity 0.25s ease, visibility 0s linear 0s; }
  /* [CHANGE 7] Panel animation — always display:flex; visibility+opacity+transform
     control show/hide. transform-origin:bottom right makes it emerge from the FAB.
     Opening: visibility fires at 0s delay so content is immediately accessible,
     then opacity/scale ease in over 0.25s.
     Closing: opacity/scale ease out first, visibility hides only after 0.25s.    */
  #ai-panel { position: fixed; right: 24px; top: calc(var(--header-h, 130px) + 0px); bottom: 85px; width: 450px; max-width: calc(100vw - 32px); background: var(--cream); border-radius: var(--radius); box-shadow: 0 12px 48px rgba(10,39,71,0.28); display: flex; flex-direction: column; overflow: hidden; z-index: 300; border: 1px solid var(--border);
    visibility: hidden; opacity: 0; transform: scale(0.90) translateY(14px); transform-origin: bottom right; pointer-events: none;
    transition: opacity 0.25s ease, transform 0.25s ease, visibility 0s linear 0.25s; }
  #ai-panel.open { visibility: visible; opacity: 1; transform: scale(1) translateY(0); pointer-events: auto;
    transition: opacity 0.25s ease, transform 0.25s ease, visibility 0s linear 0s; }
  .ai-header { background: linear-gradient(135deg, var(--navy) 0%, var(--navy-mid) 100%); color: white; padding: 14px 16px; display: flex; align-items: center; gap: 10px; flex-shrink: 0; }
  .ai-header .ai-title { font-size: 14px; font-weight: 700; }
  .ai-header .ai-sub { font-size: 10px; color: rgba(255,255,255,0.6); margin-top: 1px; }
  .ai-header-text { flex: 1; }
  .ai-close { background: none; border: none; color: rgba(255,255,255,0.7); cursor: pointer; padding: 4px; display: flex; }
  .ai-close:hover { color: white; }
  .ai-messages { flex: 1; overflow-y: auto; padding: 14px; display: flex; flex-direction: column; gap: 10px; }
  .ai-msg { max-width: 88%; font-size: 13px; line-height: 1.45; padding: 9px 12px; border-radius: 12px; white-space: pre-wrap; }
  .ai-msg.user { align-self: flex-end; background: var(--navy); color: white; border-bottom-right-radius: 3px; }
  .ai-msg.bot { align-self: flex-start; background: white; color: var(--text); border: 1px solid var(--border); border-bottom-left-radius: 3px; }
  .ai-msg.error { align-self: flex-start; background: #FEF2F2; color: #991B1B; border: 1px solid #FECACA; }
  .ai-results { display: flex; flex-direction: column; gap: 8px; align-self: stretch; }
  .ai-result-card { background: white; border: 1px solid var(--border); border-radius: 8px; padding: 10px 12px; cursor: pointer; transition: border-color 0.15s, background 0.15s; }
  .ai-result-card:hover { border-color: var(--gold); background: var(--tag-bg); }
  .ai-result-eq { font-size: 13px; font-weight: 700; color: var(--navy); margin-bottom: 2px; }
  .ai-result-lab { font-size: 11px; color: var(--text-muted); margin-bottom: 4px; }
  .ai-result-why { font-size: 11px; color: #946715; font-style: italic; }
  .ai-typing { display: flex; gap: 4px; padding: 9px 12px; align-self: flex-start; }
  .ai-typing span { width: 6px; height: 6px; border-radius: 50%; background: var(--text-muted); opacity: 0.5; animation: ai-bounce 1.2s infinite; }
  .ai-typing span:nth-child(2) { animation-delay: 0.15s; }
  .ai-typing span:nth-child(3) { animation-delay: 0.3s; }
  @keyframes ai-bounce { 0%,60%,100%{transform:translateY(0);opacity:.5} 30%{transform:translateY(-4px);opacity:1} }
  .ai-input-row { display: flex; gap: 8px; padding: 12px; border-top: 1px solid var(--border); background: white; flex-shrink: 0; }
  .ai-input-row input { flex: 1; border: 1px solid var(--border); border-radius: 8px; padding: 9px 12px; font-size: 13px; outline: none; }
  .ai-input-row input:focus { border-color: var(--gold); }
  .ai-input-row button { background: var(--navy); color: white; border: none; border-radius: 8px; width: 38px; flex-shrink: 0; cursor: pointer; display: flex; align-items: center; justify-content: center; }
  .ai-input-row button:disabled { opacity: 0.5; cursor: default; }
  /* [CHANGE 6] Disclaimer font enlarged for readability */
  .ai-disclaimer { font-size: 12px; color: var(--text-muted); padding: 8px 14px 12px; text-align: center; flex-shrink: 0; line-height: 1.5; }
  .ai-disclaimer a { color: var(--gold); }
  @media (max-width: 700px) { #ai-panel { right: 12px; left: 12px; width: auto; top: calc(var(--header-h, 130px) + 6px); bottom: 80px; } #ai-fab { right: 16px; bottom: 16px; } }
</style>
</head>
<body>
<header>
  <div class="header-top">
    <div class="logo-block">

      <!-- ── CHANGE 1: Institution logos ────────────────────────────────────────
           Two real logo images replace the old placeholder gold circle.
           Ensure iitm_logo.png and iitmrp_logo.png are in the same folder
           as the generated HTML file before opening it in a browser.
           ─────────────────────────────────────────────────────────────────── -->
      <div class="logos-wrap">
        <img src="iitm_logo.png" alt="IIT Madras" class="logo-img">
        <img src="iitmrp_logo.png" alt="IIT Madras Research Park" class="logo-img">
      </div>
      <!-- ── END CHANGE 1 ─────────────────────────────────────────────────── -->

      <div class="logo-text">
        <h1>IIT Madras/IITM Research Park Ecosystem(Beta) — Labs &amp; Equipments</h1>
        <p>Research Facilities Directory</p>
      </div>
    </div>
    <div class="header-stats">
      <div class="stat-pill"><span class="num" id="stat-total">-</span><span class="lbl">Equipment</span></div>
      <div class="stat-pill"><span class="num" id="stat-depts">-</span><span class="lbl">Departments</span></div>
      <div class="stat-pill"><span class="num" id="stat-labs">-</span><span class="lbl">Laboratories</span></div>
    </div>
  </div>
  <!-- [CHANGE 6] Contact guidance info bar -->
  <div class="header-info-bar">
    For additional details, please reach out to us at <a href="mailto:rct@respark.iitm.ac.in">rct@respark.iitm.ac.in</a>. If PI name/number is available you can contact the PI. Please prefix <span class="link-color">2257</span> before the 4 digit extension number.
  </div>
  <div class="filter-bar">

    <!-- Global search: matches equipment name, lab, professor, department, operator -->
    <div class="search-wrap">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
      <input type="text" id="search-input" placeholder="Search equipment, lab, or professor…" autocomplete="off">
    </div>

    <!-- ── CHANGE 2: Equipment-name-only search bar ───────────────────────────
         This bar filters ONLY by the equipment name field.
         Gold-tinted icon (with a + inside the magnifier) distinguishes it from
         the global search bar. Both bars can be used simultaneously.
         ─────────────────────────────────────────────────────────────────────── -->
    <div class="search-wrap equip-only-wrap">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/><path d="M8 11h6M11 8v6"/></svg>
      <input type="text" id="equip-search-input" placeholder="Search equipment name only…" autocomplete="off">
    </div>
    <!-- ── END CHANGE 2 ───────────────────────────────────────────────────── -->

    <select id="dept-filter"><option value="">All Departments</option></select>

    <!-- ════════════════════════════════════════════════════════════════════════
         CHANGE 3: ALL LABORATORIES FILTER — COMMENTED OUT
         Reason  : Not needed at this time.
         Re-enable: Remove the HTML comment tags around the <select> below,
                    then un-comment every JS block labelled "CHANGE 3 RE-ENABLE"
                    in the <script> section at the bottom of this file.
         ════════════════════════════════════════════════════════════════════════

    <select id="lab-filter"><option value="">All Laboratories</option></select>

         ════════════════════════════════════════════════════════════════════════ -->

    <button class="clear-btn" id="clear-btn">✕ Clear</button>
    <div class="view-toggle">
      <button class="view-btn active" id="btn-card" title="Card view">
        <svg width="16" height="16" fill="currentColor" viewBox="0 0 24 24"><rect x="2" y="2" width="9" height="9" rx="1"/><rect x="13" y="2" width="9" height="9" rx="1"/><rect x="2" y="13" width="9" height="9" rx="1"/><rect x="13" y="13" width="9" height="9" rx="1"/></svg>
      </button>
      <button class="view-btn" id="btn-table" title="Table view">
        <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M3 5h18M3 10h18M3 15h18M3 20h18M8 5v15M16 5v15"/></svg>
      </button>
    </div>
  </div>
</header>
<main>
  <div class="results-meta"><div class="results-count">Showing <strong id="result-count">0</strong> results</div></div>
  <div id="card-view" class="active"></div>
  <div id="table-view">
      <div class="table-scroll">
        <table>
          <thead><tr>
            <th onclick="sortTable(0)">Equipment <span class="sort-icon">↕</span></th>
            <th onclick="sortTable(1)">Laboratory <span class="sort-icon">↕</span></th>
            <th onclick="sortTable(2)">Department <span class="sort-icon">↕</span></th>
            <th onclick="sortTable(3)">Professor Incharge <span class="sort-icon">↕</span></th>
            <th>Operator</th><th>Contact</th>
          </tr></thead>
          <tbody id="table-body"></tbody>
        </table>
      </div>
    </div>
  <div class="empty-state" id="empty-state" style="display:none">
    <h3>No equipment found</h3><p>Try adjusting your search or filters.</p>
  </div>
</main>
<footer>IIT Madras Research Facilities Directory &nbsp;·&nbsp; Generated: __DATE__</footer>

<!-- ════════════════════════════════════════════════════════════════════════
     [CHANGE 5] AI Equipment Assistant — chat widget markup
     [CHANGE 6] Backdrop overlay div for dimming the background
     ════════════════════════════════════════════════════════════════════════ -->
<div id="ai-backdrop"></div>
<button id="ai-fab" title="Ask the Equipment Assistant" aria-label="Open equipment assistant">
  <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2" stroke-linecap="round" stroke-linejoin="round"><path d="M21 11.5a8.38 8.38 0 0 1-.9 3.8 8.5 8.5 0 0 1-7.6 4.7 8.38 8.38 0 0 1-3.8-.9L3 21l1.9-5.7a8.38 8.38 0 0 1-.9-3.8 8.5 8.5 0 0 1 4.7-7.6 8.38 8.38 0 0 1 3.8-.9h.5a8.48 8.48 0 0 1 8 8v.5z"/></svg>
</button>
<div id="ai-panel">
  <div class="ai-header">
    <div class="ai-header-text">
      <div class="ai-title">AI-Powered Equipment Suggestion Assistant</div>
      <div class="ai-sub">Describe your research need</div>
    </div>
    <button class="ai-close" id="ai-close" aria-label="Close assistant">
      <svg width="18" height="18" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><line x1="18" y1="6" x2="6" y2="18"/><line x1="6" y1="6" x2="18" y2="18"/></svg>
    </button>
  </div>
  <div class="ai-messages" id="ai-messages"></div>
  <div class="ai-input-row">
    <input type="text" id="ai-input" placeholder="e.g. measuring thermal conductivity of a nanofluid" autocomplete="off">
    <button id="ai-send" aria-label="Send">
      <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><line x1="22" y1="2" x2="11" y2="13"/><polygon points="22 2 15 22 11 13 2 9 22 2"/></svg>
    </button>
  </div>
  <div class="ai-disclaimer">AI-generated suggestions — please confirm availability with the listed contact or the Research Collaboration Team at <a href="mailto:rct@respark.iitm.ac.in">rct@respark.iitm.ac.in</a>.</div>
</div>

<script>
const equipmentData = __DATA__;

/* ════════════════════════════════════════════════════════════════════════
   [CHANGE 5] (JS): AI Equipment Assistant — config + candidate index
   No API key lives in this file. ASSISTANT_PROXY_URL points at a Cloudflare
   Worker (see worker.js / DEPLOYMENT.md) that holds the Gemini key
   server-side. This URL is not a secret — it's safe to commit publicly.
   ════════════════════════════════════════════════════════════════════════ */
const ASSISTANT_PROXY_URL = "__ASSISTANT_PROXY_URL__";

// Compact "index|equipment|lab|department|professor" lines — far cheaper in
// tokens than re-sending full JSON objects, while still giving the model
// everything it needs to judge relevance and to cite a valid index back.
const aiCandidateIndex = equipmentData.map((e,i)=>{
  const trunc=(s,n)=>{s=(s||'').trim();return s.length>n?s.slice(0,n)+'…':s;};
  const prof=(e.professor||'').replace(/\s*\([^)]*\)/g,'').trim();
  return `${i}|${trunc(e.equipment,90)}|${trunc(e.lab,70)}|${e.department}|${trunc(prof,40)}`;
}).join('\n');
const deptClass=(dept)=>{const d=(dept||'').toLowerCase();if(d.includes('physics'))return{card:'dept-Physics',badge:'badge-Physics'};if(d.includes('aerospace'))return{card:'dept-Aerospace-Engineering',badge:'badge-Aerospace'};if(d.includes('mechanical'))return{card:'dept-Mechanical-Engineering',badge:'badge-Mechanical'};if(d.includes('civil'))return{card:'dept-Civil-Engineering',badge:'badge-Civil'};if(d.includes('electrical'))return{card:'dept-Electrical-Engineering',badge:'badge-Electrical'};if(d.includes('chemical'))return{card:'dept-Chemical-Engineering',badge:'badge-Chemical'};if(d.includes('biotech'))return{card:'dept-Biotechnology',badge:'badge-Biotechnology'};if(d.includes('chemistry'))return{card:'dept-Chemistry',badge:'badge-Chemistry'};if(d.includes('ocean'))return{card:'dept-Ocean-Engineering',badge:'badge-Ocean'};if(d.includes('metallurg')||d.includes('material'))return{card:'dept-Metallurgical',badge:'badge-Metal'};if(d.includes('applied'))return{card:'dept-Applied-Mechanics',badge:'badge-Applied'};if(d.includes('arci'))return{card:'dept-ARCI',badge:'badge-ARCI'};if(d.includes('ic')||d.includes('icsr'))return{card:'dept-IC-SR',badge:'badge-ICSR'};if(d.includes('engineering design'))return{card:'dept-Engineering-Design',badge:'badge-ED'};return{card:'dept-default',badge:'badge-default'};};
const depts=[...new Set(equipmentData.map(e=>e.department).filter(Boolean))].sort();
const deptSel=document.getElementById('dept-filter');
depts.forEach(d=>{const o=document.createElement('option');o.value=d;o.textContent=d;deptSel.appendChild(o);});

/* ════════════════════════════════════════════════════════════════════════════
   CHANGE 3 (JS): populateLabs — FUNCTION COMMENTED OUT
   The lab-filter <select> in the HTML is also commented out (see filter-bar).
   ─────────────────────────────────────────────────────────────────────────
   CHANGE 3 RE-ENABLE ↓ (un-comment the two lines below to restore lab filter)
function populateLabs(fd){const src=fd?equipmentData.filter(e=>e.department===fd).map(e=>e.lab):equipmentData.map(e=>e.lab);const labs=[...new Set(src.filter(Boolean))].sort();const ls=document.getElementById('lab-filter');ls.innerHTML='<option value="">All Laboratories</option>';labs.forEach(l=>{const o=document.createElement('option');o.value=l;o.textContent=l.length>50?l.slice(0,50)+'…':l;ls.appendChild(o);});}
populateLabs();
   CHANGE 3 RE-ENABLE ↑
   ════════════════════════════════════════════════════════════════════════════ */

let viewMode='card',sortCol=-1,sortDir=1,currentData=[];

/* CHANGE 2 + CHANGE 3 (JS): getF() updated
   ● Added 'eq' — value from the new equipment-name-only search input (Change 2)
   ● 'lab' is hardcoded to '' — lab-filter is disabled (Change 3)
     CHANGE 3 RE-ENABLE: replace  lab:''  with  lab:document.getElementById('lab-filter').value  */
function getF(){return{q:document.getElementById('search-input').value.trim().toLowerCase(),eq:document.getElementById('equip-search-input').value.trim().toLowerCase(),dept:deptSel.value,lab:''};}

/* CHANGE 2 (JS): filterData() — added equipment-name-only check using 'eq' */
function filterData(){const{q,eq,dept,lab}=getF();return equipmentData.filter(e=>{if(dept&&e.department!==dept)return false;if(lab&&e.lab!==lab)return false;if(q){const h=[e.equipment,e.lab,e.professor,e.department,e.operator].join(' ').toLowerCase();if(!h.includes(q))return false;}if(eq&&!e.equipment.toLowerCase().includes(eq))return false;return true;});}

function esc(s){return String(s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');}
function hl(t,q){if(!q||!t)return esc(t||'');const re=new RegExp('('+q.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi');return esc(t).replace(re,'<mark style="background:#FEF08A;padding:0 1px;border-radius:2px">$1</mark>');}
function renderCards(data,q){const w=document.getElementById('card-view');if(!data.length){w.innerHTML='';return;}w.innerHTML=data.map(e=>{const cls=deptClass(e.department);const pd=e.professor||'';const em=pd.match(/\(([^)]+@[^)]+)\)/);const email=em?em[1].trim():(e.prof_email||'');const pn=pd.replace(/\s*\([^)]+\)/g,'').trim();return`<div class="eq-card ${cls.card}"><span class="card-dept-tag">${esc(e.department)}</span><div class="card-equip-name">${hl(e.equipment,q)}</div><div class="card-lab-name"><svg width="12" height="12" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>${hl(e.lab,q)}</div>${(pn||email||e.operator||e.contact)?'<div class="card-divider"></div>':''}${pn?`<div class="card-info-row"><svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg><span><span class="label">PI:</span> ${hl(pn,q)}${email?` &nbsp;<a href="mailto:${esc(email)}" class="contact-link">${esc(email)}</a>`:''}</span></div>`:''}${e.operator?`<div class="card-info-row"><svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M20 21v-2a4 4 0 0 0-4-4H8a4 4 0 0 0-4 4v2"/><circle cx="12" cy="7" r="4"/></svg><span><span class="label">Operator:</span> ${hl(e.operator,q)}${e.op_email?` &nbsp;<a href="mailto:${esc(e.op_email)}" class="contact-link">${esc(e.op_email)}</a>`:''}</span></div>`:''}${e.contact&&e.contact!==' '?`<div class="card-info-row"><svg width="13" height="13" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><path d="M22 16.92v3a2 2 0 0 1-2.18 2 19.79 19.79 0 0 1-8.63-3.07A19.5 19.5 0 0 1 4.69 11.6 19.79 19.79 0 0 1 1.6 3.08 2 2 0 0 1 3.56 1h3a2 2 0 0 1 2 1.72c.127.96.361 1.903.7 2.81a2 2 0 0 1-.45 2.11L7.91 8.69a16 16 0 0 0 5.89 5.89l.9-.9a2 2 0 0 1 2.11-.45c.907.339 1.85.573 2.81.7A2 2 0 0 1 22 16.92z"/></svg><span><span class="label">Contact:</span> ${esc(e.contact)}</span></div>`:''}</div>`;}).join('');}
function renderTable(data,q){const tb=document.getElementById('table-body');tb.innerHTML=data.map(e=>{const cls=deptClass(e.department);const pd=e.professor||'';const em=pd.match(/\(([^)]+@[^)]+)\)/);const email=em?em[1].trim():(e.prof_email||'');const pn=pd.replace(/\s*\([^)]+\)/g,'').trim();return`<tr><td>${hl(e.equipment,q)}</td><td style="color:#5A6A7A">${hl(e.lab,q)}</td><td><span class="dept-badge ${cls.badge}">${esc(e.department)}</span></td><td>${pn?hl(pn,q):''}${email?`<br><a href="mailto:${esc(email)}" class="contact-link">${esc(email)}</a>`:''}</td><td>${e.operator?hl(e.operator,q):''}${e.op_email?`<br><a href="mailto:${esc(e.op_email)}" class="contact-link">${esc(e.op_email)}</a>`:''}</td><td>${e.contact&&e.contact!==' '?esc(e.contact):''}</td></tr>`;}).join('');}

/* CHANGE 2 (JS): render() — uses q||eq as combined highlight term so both
   global-search and equipment-only-search terms are visually highlighted       */
function render(){const data=filterData();currentData=data;const{q,eq}=getF();const hlq=q||eq;document.getElementById('result-count').textContent=data.length.toLocaleString();document.getElementById('empty-state').style.display=data.length?'none':'block';if(viewMode==='card')renderCards(data,hlq);else renderTable(data,hlq);}

/* CHANGE 2 (JS): sortTable() — uses q||eq for highlight term (same as render) */
function sortTable(col){const keys=['equipment','lab','department','professor'];if(sortCol===col)sortDir*=-1;else{sortCol=col;sortDir=1;}document.querySelectorAll('th').forEach((th,i)=>{th.classList.toggle('sorted',i===col);const ic=th.querySelector('.sort-icon');if(ic)ic.textContent=i===col?(sortDir===1?'↑':'↓'):'↕';});currentData.sort((a,b)=>{const av=(a[keys[col]]||'').toLowerCase();const bv=(b[keys[col]]||'').toLowerCase();return av<bv?-sortDir:av>bv?sortDir:0;});const{q,eq}=getF();renderTable(currentData,q||eq);}

let dt;
document.getElementById('search-input').addEventListener('input',()=>{clearTimeout(dt);dt=setTimeout(render,180);});

/* CHANGE 2 (JS): equipment-name-only search input event listener */
document.getElementById('equip-search-input').addEventListener('input',()=>{clearTimeout(dt);dt=setTimeout(render,180);});

/* CHANGE 3 (JS): dept-filter change — populateLabs() and lab-filter reset removed
   CHANGE 3 RE-ENABLE: restore to:
   document.getElementById('dept-filter').addEventListener('change',()=>{populateLabs(deptSel.value);document.getElementById('lab-filter').value='';render();}); */
document.getElementById('dept-filter').addEventListener('change',()=>{render();});

/* CHANGE 3 (JS): lab-filter change listener — COMMENTED OUT
   CHANGE 3 RE-ENABLE: document.getElementById('lab-filter').addEventListener('change',render); */

/* CHANGE 2 + CHANGE 3 (JS): clear button — clears both search bars; lab-filter
   clear is removed since that dropdown is disabled (Change 3)
   CHANGE 3 RE-ENABLE: add  document.getElementById('lab-filter').value='';populateLabs();  */
document.getElementById('clear-btn').addEventListener('click',()=>{document.getElementById('search-input').value='';document.getElementById('equip-search-input').value='';deptSel.value='';render();});

/* CHANGE 4 (JS): view toggle buttons — updated to pass q||eq as highlight term */
document.getElementById('btn-card').addEventListener('click',()=>{viewMode='card';document.getElementById('btn-card').classList.add('active');document.getElementById('btn-table').classList.remove('active');document.getElementById('card-view').classList.add('active');document.getElementById('table-view').classList.remove('active');const{q,eq}=getF();renderCards(currentData,q||eq);});
document.getElementById('btn-table').addEventListener('click',()=>{viewMode='table';document.getElementById('btn-table').classList.add('active');document.getElementById('btn-card').classList.remove('active');document.getElementById('table-view').classList.add('active');document.getElementById('card-view').classList.remove('active');const{q,eq}=getF();renderTable(currentData,q||eq);});

document.getElementById('stat-total').textContent=equipmentData.length.toLocaleString();
document.getElementById('stat-depts').textContent=depts.length;
document.getElementById('stat-labs').textContent=[...new Set(equipmentData.map(e=>e.lab).filter(Boolean))].length;

/* CHANGE 4 (JS): Compute and store the page header height as a CSS variable.
   --header-h is used by  thead th { top: var(--header-h) }  so the sticky
   table header sits flush below the sticky page header rather than behind it.
   ResizeObserver keeps the value current if the filter-bar wraps on resize.    */
function updateHeaderHeight(){const h=document.querySelector('header');if(h)document.documentElement.style.setProperty('--header-h',h.offsetHeight+'px');}
updateHeaderHeight();
if(window.ResizeObserver){new ResizeObserver(updateHeaderHeight).observe(document.querySelector('header'));}
else{window.addEventListener('resize',updateHeaderHeight);}

/* ════════════════════════════════════════════════════════════════════════
   [CHANGE 5] (JS): AI Equipment Assistant — chat logic
   ════════════════════════════════════════════════════════════════════════ */
const AI_SYSTEM_PROMPT = `You are an equipment-matching assistant for the IIT Madras Research Park Labs & Equipment Directory. The user describes a research idea, technique, or measurement need in their own words — it may be vague, casual, or use different terminology than the equipment names. Your job is to infer the underlying technique or instrument family and select the most relevant entries from the CANDIDATE LIST given to you (format per line: index|equipment|lab|department|professor).

Rules:
- Only use indices that literally appear in the candidate list. Never invent equipment, labs, or indices.
- Rank by genuine relevance to the stated need, not by surface keyword overlap.
- Return at most 8 matches. Return fewer, or none, if fewer are genuinely relevant — do not pad the list.
- If nothing fits, return an empty "matches" array and use "reply" to suggest what kind of facility might help, or ask one clarifying question.
- "reply" must be 1-2 short, conversational sentences. No markdown, no headers, no lists inside "reply".
- "why" per match must be <=12 words, plain language.
- Respond with ONLY valid JSON, exactly this shape, no code fences, no extra text:
{"reply":"string","matches":[{"i":0,"why":"string"}]}`;

let aiHistory = [];          // [{role:'user'|'model', parts:[{text}]}]
let aiCandidatesSent = false; // candidate list is sent once, then lives in history
let aiBusy = false;

// [CHANGE 7] Two flags that track the session's message lifecycle:
//   aiWelcomeShown  — true once the first-open greeting has been rendered;
//                     prevents the welcome from re-appearing on every open.
//   aiHasInteracted — true after the user's first sent message;
//                     gates whether a re-open continuation prompt is warranted.
let aiWelcomeShown  = false;
let aiHasInteracted = false;

// Professionally-varied continuation prompts — one is selected at random on
// each re-open so the prompt never feels mechanical or repetitive.
const AI_CONTINUATION_PROMPTS = [
  "Welcome back — feel free to refine your previous search or explore a new requirement.",
  "Your prior conversation is intact. Is there a follow-up query I can assist with?",
  "Ready to continue. Would you like to narrow the results further or try a different research need?",
  "Happy to help further — let me know if you'd like to adjust your criteria or explore another area."
];

function aiOpen(){
  const panel    = document.getElementById('ai-panel');
  const backdrop = document.getElementById('ai-backdrop');
  const isReopen = aiWelcomeShown && !panel.classList.contains('open');

  panel.classList.add('open');
  backdrop.classList.add('show');

  if (!aiWelcomeShown) {
    // First-ever open: show the welcome greeting exactly once.
    aiAppendBot("Hi! I\u2019m your IITMRP Equipment Suggestion Bot. Describe what you\u2019re trying to build, test, or measure and I\u2019ll find the most relevant labs and equipment for you.");
    aiWelcomeShown = true;
  } else if (isReopen && aiHasInteracted) {
    // Subsequent re-opens after the user has interacted: surface a brief,
    // professional continuation prompt so the conversation feels alive.
    const prompt = AI_CONTINUATION_PROMPTS[Math.floor(Math.random() * AI_CONTINUATION_PROMPTS.length)];
    aiAppendBot(prompt);
  }
  // If the user closes and re-opens before ever sending a message, nothing
  // extra is appended — the existing welcome remains the last message.

  document.getElementById('ai-input').focus();
}
function aiClose(){
  document.getElementById('ai-panel').classList.remove('open');
  document.getElementById('ai-backdrop').classList.remove('show');
}

function aiAppendUser(text){
  const w=document.getElementById('ai-messages');
  const d=document.createElement('div'); d.className='ai-msg user'; d.textContent=text;
  w.appendChild(d); w.scrollTop=w.scrollHeight;
}
function aiAppendBot(text, isError){
  const w=document.getElementById('ai-messages');
  const d=document.createElement('div'); d.className='ai-msg bot'+(isError?' error':''); d.textContent=text;
  w.appendChild(d); w.scrollTop=w.scrollHeight;
}
function aiAppendResults(matches){
  const w=document.getElementById('ai-messages');
  const wrap=document.createElement('div'); wrap.className='ai-results';
  matches.forEach(m=>{
    const e=equipmentData[m.i]; if(!e) return;
    const card=document.createElement('div'); card.className='ai-result-card';
    card.innerHTML=`<div class="ai-result-eq">${esc(e.equipment)}</div><div class="ai-result-lab">${esc(e.lab)} · ${esc(e.department)}</div>${m.why?`<div class="ai-result-why">${esc(m.why)}</div>`:''}`;
    card.addEventListener('click', ()=>aiJumpTo(e));
    wrap.appendChild(card);
  });
  if(wrap.children.length){ w.appendChild(wrap); w.scrollTop=w.scrollHeight; }
}
function aiJumpTo(e){
  document.getElementById('equip-search-input').value='';
  document.getElementById('search-input').value=e.equipment;
  deptSel.value='';
  render();
  aiClose();
  document.querySelector('main').scrollIntoView({behavior:'smooth'});
}
function aiTypingShow(){
  const w=document.getElementById('ai-messages');
  const d=document.createElement('div'); d.className='ai-typing'; d.id='ai-typing';
  d.innerHTML='<span></span><span></span><span></span>';
  w.appendChild(d); w.scrollTop=w.scrollHeight;
}
function aiTypingHide(){ const t=document.getElementById('ai-typing'); if(t) t.remove(); }

async function aiSend(){
  if(aiBusy) return;
  const input=document.getElementById('ai-input');
  const text=input.value.trim();
  if(!text) return;
  if(!ASSISTANT_PROXY_URL){
    aiAppendUser(text); input.value='';
    aiAppendBot("The assistant isn't configured yet — deploy the proxy (see DEPLOYMENT.md) and set ASSISTANT_PROXY_URL in regenerate_browser.py, then regenerate this page.", true);
    return;
  }
  aiBusy=true;
  document.getElementById('ai-send').disabled=true;
  aiHasInteracted = true; // [CHANGE 7] gates continuation prompt on next re-open
  aiAppendUser(text); input.value='';

  const includeCandidates = !aiCandidatesSent;
  const userMessage = includeCandidates
    ? `CANDIDATE LIST:\n${aiCandidateIndex}\n\nUSER REQUEST: ${text}`
    : text;
  aiHistory.push({role:'user', parts:[{text:userMessage}]});
  aiTypingShow();

  try{
    const res = await fetch(ASSISTANT_PROXY_URL, {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({
        systemInstruction: { parts:[{text:AI_SYSTEM_PROMPT}] },
        contents: aiHistory,
        generationConfig: { responseMimeType:'application/json', temperature:0.3, thinkingConfig:{thinkingBudget:0} }
      })
    });
    aiTypingHide();

    if(!res.ok){
      let msg = 'Something went wrong reaching the assistant. Please try again.';
      if(res.status===429) msg = 'The assistant has hit its free-tier usage limit for now — please try again in a minute.';
      else if(res.status===403) msg = "The assistant proxy rejected this request (origin not allowed). If you just deployed it, check ALLOWED_ORIGINS in worker.js matches this page's URL.";
      else if(res.status===400) msg = 'The assistant proxy is reachable but the request was rejected — check the Worker logs.';
      aiAppendBot(msg, true);
      aiHistory.pop();
      return;
    }

    const data = await res.json();
    // Gemini 2.5 Flash can return multiple parts when thinking is active.
    // thought:true marks internal reasoning — skip those and find the real output.
    const allParts = data?.candidates?.[0]?.content?.parts || [];
    const outputPart = allParts.find(p => !p.thought) || allParts[0] || {};
    const raw = outputPart.text || '';
    aiHistory.push({role:'model', parts:[{text:raw}]});
    if(includeCandidates) aiCandidatesSent = true;

    const MAX_HISTORY_TURNS = 10;
    if(aiHistory.length > MAX_HISTORY_TURNS * 2){
      aiHistory = aiHistory.slice(-(MAX_HISTORY_TURNS * 2));
    }

    let parsed;
    try{
      parsed = JSON.parse(raw.replace(/```json|```/g,'').trim());
    }catch(e){
      aiAppendBot(raw || "I couldn't parse a response — please try rephrasing.", true);
      return;
    }
    aiAppendBot(parsed.reply || "Here's what I found:");
    const matches=(parsed.matches||[]).filter(m=>Number.isInteger(m.i) && m.i>=0 && m.i<equipmentData.length).slice(0,8);
    if(matches.length) aiAppendResults(matches);
  }catch(err){
    aiTypingHide();
    aiAppendBot('Network error reaching the assistant. Check your connection and try again.', true);
    aiHistory.pop();
  }finally{
    aiBusy=false;
    document.getElementById('ai-send').disabled=false;
    input.focus();
  }
}

document.getElementById('ai-fab').addEventListener('click', aiOpen);
document.getElementById('ai-close').addEventListener('click', aiClose);
document.getElementById('ai-backdrop').addEventListener('click', aiClose); // click backdrop to close
document.getElementById('ai-send').addEventListener('click', aiSend);
document.getElementById('ai-input').addEventListener('keydown', (e)=>{ if(e.key==='Enter') aiSend(); });

render();

// [CHANGE 6] Auto-open the assistant when the page first loads
aiOpen();
</script>
</body>
</html>"""


def generate_html(equipment, output_path):
    from datetime import date
    data_json = json.dumps(equipment, separators=(',', ':'))
    html = HTML_TEMPLATE.replace('__DATA__', data_json)
    html = html.replace('__DATE__', date.today().strftime('%B %d, %Y'))
    html = html.replace('__ASSISTANT_PROXY_URL__', ASSISTANT_PROXY_URL)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Written: {output_path}  ({len(html):,} bytes)")


if __name__ == '__main__':
    base = os.path.dirname(os.path.abspath(__file__))
    source = os.path.join(base, SOURCE_FILE)
    output = os.path.join(base, OUTPUT_FILE)

    if not os.path.exists(source):
        print(f"ERROR: Source file not found: {source}")
        print("Place this script in the same folder as the Excel file.")
        exit(1)

    print(f"Reading: {SOURCE_FILE}")
    equipment = extract_data(source)

    print(f"Generating HTML…")
    generate_html(equipment, output)

    print(f"\n✓ Done! Open {OUTPUT_FILE} in any browser.")
    print(f"  {len(equipment)} equipment entries across "
          f"{len(set(e['department'] for e in equipment))} departments.")

    if ASSISTANT_PROXY_URL:
        print(f"\n  AI Equipment Assistant is wired to: {ASSISTANT_PROXY_URL}")
        print("    No API key is embedded in this HTML — confirm the Worker is")
        print("    deployed and ALLOWED_ORIGINS in worker.js includes the exact")
        print("    URL this page will be hosted at (e.g. your github.io URL).")
    else:
        print("\n  ℹ AI Equipment Assistant is disabled (ASSISTANT_PROXY_URL is empty).")
        print("    See DEPLOYMENT.md to deploy the free Cloudflare Worker proxy,")
        print("    then paste its URL into ASSISTANT_PROXY_URL near the top of")
        print("    this script and re-run. Do NOT put a Gemini API key directly")
        print("    in this script or in the generated HTML — it would be")
        print("    committed to your repo and exposed to every site visitor.")
